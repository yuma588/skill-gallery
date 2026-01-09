"""创建项目进度跟踪电子表格

该脚本创建一个完整的项目管理Excel文件,包含:
- 项目概览: 显示关键指标和统计信息
- 任务列表: 详细的项目任务跟踪
- 团队信息: 团队成员和工作分配
"""

from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule


def create_header_style():
    """创建标题样式"""
    return Font(
        name='微软雅黑',
        size=12,
        bold=True,
        color='FFFFFF'
    ), PatternFill(
        start_color='4472C4',
        end_color='4472C4',
        fill_type='solid'
    ), Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )


def create_subheader_style():
    """创建副标题样式"""
    return Font(
        name='微软雅黑',
        size=11,
        bold=True,
        color='FFFFFF'
    ), PatternFill(
        start_color='5B9BD5',
        end_color='5B9BD5',
        fill_type='solid'
    ), Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )


def create_thin_border():
    """创建细边框"""
    thin = Side(border_style='thin', color='000000')
    return Border(left=thin, top=thin, right=thin, bottom=thin)


def create_project_overview_sheet(wb):
    """创建项目概览工作表"""
    ws = wb.active
    ws.title = '项目概览'

    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # 获取样式
    header_font, header_fill, header_alignment = create_header_style()
    subheader_font, subheader_fill, subheader_alignment = create_subheader_style()
    border = create_thin_border()

    # 标题
    ws['A1'] = '📊 项目进度跟踪系统'
    ws.merge_cells('A1:C1')
    ws['A1'].font = Font(name='微软雅黑', size=16, bold=True, color='1F4E78')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # 项目基本信息
    ws['A3'] = '项目信息'
    ws['B3'] = '项目名称'
    ws['C3'] = '数字化转型项目'
    ws['B4'] = '项目经理'
    ws['C4'] = '张三'
    ws['B5'] = '开始日期'
    ws['C5'] = '2026-01-01'
    ws['B6'] = '结束日期'
    ws['C6'] = '2026-06-30'
    ws['B7'] = '项目周期'
    ws['C7'] = '181天'

    # 应用样式到项目信息区域
    for row in range(3, 8):
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = subheader_fill
        ws[f'A{row}'].alignment = subheader_alignment
        for col in ['B', 'C']:
            cell = ws[f'{col}{row}']
            cell.border = border
            cell.font = Font(name='微软雅黑', size=10)

    # 关键指标
    ws['A10'] = '📈 关键指标'
    ws.merge_cells('A10:C10')
    ws['A10'].font = header_font
    ws['A10'].fill = header_fill
    ws['A10'].alignment = header_alignment

    # 指标表头
    headers = ['指标', '数值', '状态']
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=11, column=idx, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = subheader_alignment
        cell.border = border

    # 指标数据
    metrics = [
        ['总任务数', '24', '=COUNTA(任务列表!A2:A25)'],
        ['已完成', '8', '=COUNTIF(任务列表!H:H,"已完成")'],
        ['进行中', '10', '=COUNTIF(任务列表!H:H,"进行中")'],
        ['未开始', '5', '=COUNTIF(任务列表!H:H,"未开始")'],
        ['已延期', '1', '=COUNTIF(任务列表!H:H,"已延期")'],
        ['总体进度', '54%', '=AVERAGE(任务列表!F:F)'],
        ['高优先级', '6', '=COUNTIF(任务列表!G:G,"高")'],
        ['中优先级', '14', '=COUNTIF(任务列表!G:G,"中")'],
        ['低优先级', '4', '=COUNTIF(任务列表!G:G,"低")'],
    ]

    for idx, metric in enumerate(metrics, 12):
        for col_idx, value in enumerate(metric, 1):
            cell = ws.cell(row=idx, column=col_idx, value=value)
            cell.border = border
            if col_idx == 1:
                cell.font = Font(name='微软雅黑', size=10, bold=True)
            else:
                cell.font = Font(name='微软雅黑', size=10)

    # 状态指示器
    ws['A22'] = '状态图例'
    ws.merge_cells('A22:C22')
    ws['A22'].font = header_font
    ws['A22'].fill = header_fill
    ws['A22'].alignment = header_alignment

    status_colors = [
        ('已完成', 'C6EFCE', '006100'),
        ('进行中', 'FFEB9C', '9C5700'),
        ('未开始', 'E2EFDA', '1B5E20'),
        ('已延期', 'FFC7CE', '9C0006'),
    ]

    for idx, (status, fill_color, font_color) in enumerate(status_colors, 23):
        ws[f'A{idx}'] = status
        ws[f'A{idx}'].font = Font(name='微软雅黑', size=10, bold=True, color=font_color)
        ws[f'A{idx}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        ws[f'A{idx}'].alignment = Alignment(horizontal='center', vertical='center')

    # 设置行高
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[10].height = 25
    ws.row_dimensions[22].height = 25

    # 冻结窗格
    ws.freeze_panes = 'A12'


def create_tasks_sheet(wb):
    """创建任务列表工作表"""
    ws = wb.create_sheet('任务列表')

    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 30

    # 获取样式
    header_font, header_fill, header_alignment = create_header_style()
    border = create_thin_border()

    # 标题
    ws['A1'] = '📋 任务跟踪表'
    ws.merge_cells('A1:I1')
    ws['A1'].font = Font(name='微软雅黑', size=16, bold=True, color='1F4E78')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # 表头
    headers = ['ID', '任务名称', '负责人', '开始日期', '结束日期', '进度(%)', '优先级', '状态', '备注']
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # 示例数据
    base_date = datetime(2026, 1, 1)
    tasks = [
        ['T01', '需求分析与规划', '张三', base_date.strftime('%Y-%m-%d'), (base_date + timedelta(days=14)).strftime('%Y-%m-%d'), 100, '高', '已完成', '需求文档已确认'],
        ['T02', '系统架构设计', '李四', (base_date + timedelta(days=15)).strftime('%Y-%m-%d'), (base_date + timedelta(days=30)).strftime('%Y-%m-%d'), 100, '高', '已完成', '架构评审通过'],
        ['T03', '数据库设计', '王五', (base_date + timedelta(days=16)).strftime('%Y-%m-%d'), (base_date + timedelta(days=35)).strftime('%Y-%m-%d'), 100, '高', '已完成', '数据库建模完成'],
        ['T04', '前端框架搭建', '赵六', (base_date + timedelta(days=31)).strftime('%Y-%m-%d'), (base_date + timedelta(days=50)).strftime('%Y-%m-%d'), 80, '高', '进行中', '正在开发UI组件'],
        ['T05', '后端API开发', '钱七', (base_date + timedelta(days=36)).strftime('%Y-%m-%d'), (base_date + timedelta(days=60)).strftime('%Y-%m-%d'), 70, '高', '进行中', '核心接口已完成80%'],
        ['T06', '用户认证模块', '孙八', (base_date + timedelta(days=40)).strftime('%Y-%m-%d'), (base_date + timedelta(days=55)).strftime('%Y-%m-%d'), 90, '高', '进行中', '正在测试'],
        ['T07', '数据可视化', '周九', (base_date + timedelta(days=45)).strftime('%Y-%m-%d'), (base_date + timedelta(days=70)).strftime('%Y-%m-%d'), 60, '中', '进行中', '图表组件开发中'],
        ['T08', '报表功能', '吴十', (base_date + timedelta(days=50)).strftime('%Y-%m-%d'), (base_date + timedelta(days=75)).strftime('%Y-%m-%d'), 45, '中', '进行中', '数据查询优化'],
        ['T09', '移动端适配', '郑十一', (base_date + timedelta(days=55)).strftime('%Y-%m-%d'), (base_date + timedelta(days=80)).strftime('%Y-%m-%d'), 30, '中', '进行中', '响应式设计调整'],
        ['T10', '性能优化', '王十二', (base_date + timedelta(days=60)).strftime('%Y-%m-%d'), (base_date + timedelta(days=85)).strftime('%Y-%m-%d'), 20, '中', '进行中', '代码重构进行中'],
        ['T11', '单元测试', '李十三', (base_date + timedelta(days=30)).strftime('%Y-%m-%d'), (base_date + timedelta(days=45)).strftime('%Y-%m-%d'), 100, '高', '已完成', '测试覆盖率90%'],
        ['T12', '集成测试', '赵十四', (base_date + timedelta(days=50)).strftime('%Y-%m-%d'), (base_date + timedelta(days=65)).strftime('%Y-%m-%d'), 75, '高', '进行中', '自动化测试脚本'],
        ['T13', '系统部署', '钱十五', (base_date + timedelta(days=80)).strftime('%Y-%m-%d'), (base_date + timedelta(days=90)).strftime('%Y-%m-%d'), 0, '高', '未开始', '等待测试完成'],
        ['T14', '用户培训', '孙十六', (base_date + timedelta(days=85)).strftime('%Y-%m-%d'), (base_date + timedelta(days=95)).strftime('%Y-%m-%d'), 0, '中', '未开始', '准备培训材料'],
        ['T15', '文档编写', '周十七', (base_date + timedelta(days=20)).strftime('%Y-%m-%d'), (base_date + timedelta(days=40)).strftime('%Y-%m-%d'), 100, '中', '已完成', '技术文档已归档'],
        ['T16', '安全审计', '吴十八', (base_date + timedelta(days=70)).strftime('%Y-%m-%d'), (base_date + timedelta(days=85)).strftime('%Y-%m-%d'), 0, '高', '未开始', '等待第三方'],
        ['T17', 'API文档', '郑十九', (base_date + timedelta(days=40)).strftime('%Y-%m-%d'), (base_date + timedelta(days=60)).strftime('%Y-%m-%d'), 85, '中', '进行中', 'Swagger文档完善中'],
        ['T18', 'Bug修复', '王二十', (base_date + timedelta(days=1)).strftime('%Y-%m-%d'), (base_date + timedelta(days=180)).strftime('%Y-%m-%d'), 50, '高', '进行中', '持续修复中'],
        ['T19', '备份策略', '李二十一', (base_date + timedelta(days=15)).strftime('%Y-%m-%d'), (base_date + timedelta(days=30)).strftime('%Y-%m-%d'), 100, '高', '已完成', '备份系统已上线'],
        ['T20', '监控系统', '赵二十二', (base_date + timedelta(days=25)).strftime('%Y-%m-%d'), (base_date + timedelta(days=50)).strftime('%Y-%m-%d'), 95, '中', '进行中', '告警规则配置中'],
        ['T21', '日志系统', '钱二十三', (base_date + timedelta(days=20)).strftime('%Y-%m-%d'), (base_date + timedelta(days=45)).strftime('%Y-%m-%d'), 100, '低', '已完成', 'ELK已部署'],
        ['T22', '邮件通知', '孙二十四', (base_date + timedelta(days=60)).strftime('%Y-%m-%d'), (base_date + timedelta(days=75)).strftime('%Y-%m-%d'), 40, '低', '进行中', '模板设计中'],
        ['T23', '短信通知', '周二十五', (base_date + timedelta(days=65)).strftime('%Y-%m-%d'), (base_date + timedelta(days=80)).strftime('%Y-%m-%d'), 0, '低', '未开始', '等待审批'],
        ['T24', '项目验收', '张三', (base_date + timedelta(days=85)).strftime('%Y-%m-%d'), (base_date + timedelta(days=90)).strftime('%Y-%m-%d'), 0, '高', '未开始', '准备验收材料'],
    ]

    # 填充数据
    for row_idx, task in enumerate(tasks, 3):
        for col_idx, value in enumerate(task, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.font = Font(name='微软雅黑', size=10)

            # 根据状态设置颜色
            if col_idx == 8:  # 状态列
                status = str(value)
                if status == '已完成':
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    cell.font = Font(name='微软雅黑', size=10, bold=True, color='006100')
                elif status == '进行中':
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    cell.font = Font(name='微软雅黑', size=10, bold=True, color='9C5700')
                elif status == '未开始':
                    cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                    cell.font = Font(name='微软雅黑', size=10, bold=True, color='1B5E20')
                elif status == '已延期':
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    cell.font = Font(name='微软雅黑', size=10, bold=True, color='9C0006')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 优先级颜色
            elif col_idx == 7:  # 优先级列
                priority = str(value)
                if priority == '高':
                    cell.font = Font(name='微软雅黑', size=10, bold=True, color='C00000')
                elif priority == '中':
                    cell.font = Font(name='微软雅黑', size=10, bold=True, color='ED7D31')
                elif priority == '低':
                    cell.font = Font(name='微软雅黑', size=10, bold=True, color='00B050')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 进度列格式
            elif col_idx == 6:  # 进度列
                if isinstance(value, (int, float)):
                    cell.number_format = '0%'
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 日期格式
            elif col_idx in [4, 5]:  # 开始日期和结束日期
                cell.number_format = 'YYYY-MM-DD'
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 其他列
            else:
                if col_idx not in [2, 9]:  # 不是任务名称和备注列
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 设置行高
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25

    # 冻结窗格
    ws.freeze_panes = 'A3'


def create_team_sheet(wb):
    """创建团队信息工作表"""
    ws = wb.create_sheet('团队信息')

    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 30

    # 获取样式
    header_font, header_fill, header_alignment = create_header_style()
    border = create_thin_border()

    # 标题
    ws['A1'] = '👥 团队成员信息'
    ws.merge_cells('A1:G1')
    ws['A1'].font = Font(name='微软雅黑', size=16, bold=True, color='1F4E78')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # 表头
    headers = ['姓名', '角色', '邮箱', '电话', '部门', '负责任务数', '备注']
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # 团队成员数据
    team_members = [
        ['张三', '项目经理', 'zhangsan@company.com', '138-0000-0001', '项目管理', '=COUNTIF(任务列表!C:C,"张三")', '项目负责人'],
        ['李四', '架构师', 'lisi@company.com', '138-0000-0002', '技术部', '=COUNTIF(任务列表!C:C,"李四")', '系统架构设计'],
        ['王五', '后端开发', 'wangwu@company.com', '138-0000-0003', '技术部', '=COUNTIF(任务列表!C:C,"王五")', '数据库专家'],
        ['赵六', '前端开发', 'zhaoliu@company.com', '138-0000-0004', '技术部', '=COUNTIF(任务列表!C:C,"赵六")', 'React/Vue专家'],
        ['钱七', '后端开发', 'qianqi@company.com', '138-0000-0005', '技术部', '=COUNTIF(任务列表!C:C,"钱七")', 'API开发'],
        ['孙八', '全栈开发', 'sunba@company.com', '138-0000-0006', '技术部', '=COUNTIF(任务列表!C:C,"孙八")', '认证模块'],
        ['周九', '前端开发', 'zhoujiu@company.com', '138-0000-0007', '技术部', '=COUNTIF(任务列表!C:C,"周九")', '数据可视化'],
        ['吴十', '后端开发', 'wushi@company.com', '138-0000-0008', '技术部', '=COUNTIF(任务列表!C:C,"吴十")', '报表功能'],
        ['郑十一', '移动端开发', 'zheng11@company.com', '138-0000-0009', '技术部', '=COUNTIF(任务列表!C:C,"郑十一")', '响应式设计'],
        ['王十二', '性能优化', 'wang12@company.com', '138-0000-0010', '技术部', '=COUNTIF(任务列表!C:C,"王十二")', '性能调优'],
    ]

    # 填充数据
    for row_idx, member in enumerate(team_members, 3):
        for col_idx, value in enumerate(member, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # 姓名列居中
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            # 角色列居中
            elif col_idx == 2:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # 设置行高
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25

    # 冻结窗格
    ws.freeze_panes = 'A3'


def create_progress_summary_sheet(wb):
    """创建进度汇总工作表"""
    ws = wb.create_sheet('进度汇总')

    # 设置列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20

    # 获取样式
    header_font, header_fill, header_alignment = create_header_style()
    subheader_font, subheader_fill, subheader_alignment = create_subheader_style()
    border = create_thin_border()

    # 标题
    ws['A1'] = '📊 项目进度汇总'
    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(name='微软雅黑', size=16, bold=True, color='1F4E78')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # 按阶段分类
    ws['A3'] = '按项目阶段'
    ws.merge_cells('A3:E3')
    ws['A3'].font = header_font
    ws['A3'].fill = header_fill
    ws['A3'].alignment = header_alignment

    # 阶段表头
    ws['A4'] = '阶段'
    ws['B4'] = '任务数'
    ws['C4'] = '已完成'
    ws['D4'] = '进行中'
    ws['E4'] = '进度(%)'

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}4'].font = subheader_font
        ws[f'{col}4'].fill = subheader_fill
        ws[f'{col}4'].alignment = subheader_alignment
        ws[f'{col}4'].border = border

    # 阶段数据
    phases = [
        ['规划阶段', 3, 3, 0, 100],
        ['设计阶段', 2, 2, 0, 100],
        ['开发阶段', 14, 2, 12, 58],
        ['测试阶段', 3, 1, 2, 75],
        ['部署阶段', 2, 0, 0, 0],
    ]

    for idx, phase in enumerate(phases, 5):
        for col_idx, value in enumerate(phase, 1):
            cell = ws.cell(row=idx, column=col_idx, value=value)
            cell.border = border
            cell.font = Font(name='微软雅黑', size=10)

            if col_idx == 5:  # 进度列
                cell.number_format = '0.0%'
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 按优先级分类
    ws['A11'] = '按优先级'
    ws.merge_cells('A11:E11')
    ws['A11'].font = header_font
    ws['A11'].fill = header_fill
    ws['A11'].alignment = header_alignment

    ws['A12'] = '优先级'
    ws['B12'] = '任务数'
    ws['C12'] = '已完成'
    ws['D12'] = '进行中'
    ws['E12'] = '进度(%)'

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}12'].font = subheader_font
        ws[f'{col}12'].fill = subheader_fill
        ws[f'{col}12'].alignment = subheader_alignment
        ws[f'{col}12'].border = border

    priorities = [
        ['高优先级', 8, 4, 4, 65],
        ['中优先级', 12, 3, 9, 47],
        ['低优先级', 4, 1, 1, 40],
    ]

    for idx, priority in enumerate(priorities, 13):
        for col_idx, value in enumerate(priority, 1):
            cell = ws.cell(row=idx, column=col_idx, value=value)
            cell.border = border

            if col_idx == 1:  # 优先级列
                if value == '高优先级':
                    cell.font = Font(name='微软雅黑', size=10, bold=True, color='C00000')
                elif value == '中优先级':
                    cell.font = Font(name='微软雅黑', size=10, bold=True, color='ED7D31')
                elif value == '低优先级':
                    cell.font = Font(name='微软雅黑', size=10, bold=True, color='00B050')
            else:
                cell.font = Font(name='微软雅黑', size=10)

            if col_idx == 5:  # 进度列
                cell.number_format = '0.0%'

            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 按成员分类
    ws['A17'] = '按成员工作量'
    ws.merge_cells('A17:E17')
    ws['A17'].font = header_font
    ws['A17'].fill = header_fill
    ws['A17'].alignment = header_alignment

    ws['A18'] = '成员'
    ws['B18'] = '任务数'
    ws['C18'] = '已完成'
    ws['D18'] = '进行中'
    ws['E18'] = '进度(%)'

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}18'].font = subheader_font
        ws[f'{col}18'].fill = subheader_fill
        ws[f'{col}18'].alignment = subheader_alignment
        ws[f'{col}18'].border = border

    members = [
        ['张三', 2, 2, 0, 100],
        ['李四', 1, 1, 0, 100],
        ['王五', 1, 1, 0, 100],
        ['赵六', 1, 0, 1, 80],
        ['钱七', 1, 0, 1, 70],
        ['孙八', 1, 0, 1, 90],
        ['周九', 1, 0, 1, 60],
        ['吴十', 1, 0, 1, 45],
        ['郑十一', 1, 0, 1, 30],
        ['王十二', 1, 0, 1, 20],
        ['李十三', 1, 1, 0, 100],
        ['赵十四', 1, 0, 1, 75],
        ['钱十五', 1, 0, 0, 0],
        ['孙十六', 1, 0, 0, 0],
        ['周十七', 1, 1, 0, 100],
        ['吴十八', 1, 0, 0, 0],
        ['郑十九', 1, 0, 1, 85],
        ['王二十', 1, 0, 1, 50],
        ['钱二十一', 1, 1, 0, 100],
        ['赵二十二', 1, 0, 1, 95],
        ['钱二十三', 1, 1, 0, 100],
        ['孙二十四', 1, 0, 1, 40],
        ['周二十五', 1, 0, 0, 0],
    ]

    for idx, member in enumerate(members, 19):
        for col_idx, value in enumerate(member, 1):
            cell = ws.cell(row=idx, column=col_idx, value=value)
            cell.border = border

            if col_idx == 1:  # 姓名列
                cell.font = Font(name='微软雅黑', size=10)
            else:
                cell.font = Font(name='微软雅黑', size=10)

            if col_idx == 5:  # 进度列
                cell.number_format = '0.0%'

            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 设置行高
    ws.row_dimensions[1].height = 30
    for row in [3, 11, 17]:
        ws.row_dimensions[row].height = 25

    # 冻结窗格
    ws.freeze_panes = 'A5'


def main():
    """主函数"""
    print("正在创建项目进度跟踪电子表格...")

    # 创建工作簿
    wb = openpyxl.Workbook()

    # 创建各个工作表
    create_project_overview_sheet(wb)
    create_tasks_sheet(wb)
    create_team_sheet(wb)
    create_progress_summary_sheet(wb)

    # 保存文件
    output_file = '项目进度跟踪表.xlsx'
    wb.save(output_file)

    print("[OK] 电子表格已成功创建: {}".format(output_file))
    print("\n[INFO] 包含以下工作表:")
    print("  1. 项目概览 - 显示关键指标和项目信息")
    print("  2. 任务列表 - 详细的任务跟踪和进度管理")
    print("  3. 团队信息 - 团队成员和工作分配")
    print("  4. 进度汇总 - 多维度的进度统计")
    print("\n[INFO] 特性:")
    print("  + 自动计算公式")
    print("  + 状态颜色标识")
    print("  + 优先级分类")
    print("  + 冻结窗格")
    print("  + 专业格式化")


if __name__ == '__main__':
    main()
